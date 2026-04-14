import argparse
import os
import re
import sys
from urllib.parse import quote_plus

try:
    from tqdm import tqdm
except ImportError:
    tqdm = None

try:
    import requests
except ImportError:
    requests = None
    import urllib.request
    import json

from openpyxl import Workbook, load_workbook

DEFAULT_TIMEOUT = 20


def http_get_text(url, timeout=DEFAULT_TIMEOUT):
    if requests is not None:
        response = requests.get(url, timeout=timeout, headers={"User-Agent": "Mozilla/5.0"})
        response.raise_for_status()
        return response.text

    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return response.read().decode("utf-8", errors="replace")


def http_get_json(url, timeout=DEFAULT_TIMEOUT):
    if requests is not None:
        response = requests.get(url, timeout=timeout, headers={"User-Agent": "Mozilla/5.0"})
        response.raise_for_status()
        return response.json()

    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def safe(value):
    return "" if value is None else str(value).strip()


def build_full_name(person):
    navn = (person or {}).get("navn", {})
    parts = [
        safe(navn.get("fornavn")),
        safe(navn.get("mellomnavn")),
        safe(navn.get("etternavn")),
    ]
    return " ".join([p for p in parts if p]).strip()


def extract_brreg_candidates(data):
    candidates = []

    for gruppe in data.get("rollegrupper", []) or []:
        gruppe_kode = safe((gruppe.get("type") or {}).get("kode"))

        for rolle in gruppe.get("roller", []) or []:
            if rolle.get("fratraadt") is True or rolle.get("avregistrert") is True:
                continue

            person = rolle.get("person")
            if not person:
                continue

            navn = build_full_name(person)
            if not navn:
                continue

            rolle_kode = safe((rolle.get("type") or {}).get("kode"))

            if rolle_kode == "KONT":
                candidates.append(("KONT", navn))
            elif rolle_kode == "DAGL":
                candidates.append(("DAGL", navn))
            elif rolle_kode == "LEDE" and gruppe_kode == "STYRE":
                candidates.append(("STYR", navn))

    return candidates


def extract_regnskapsforer(data):
    """Extract regnskapsfører (accountant) name from Brreg roles data."""
    for gruppe in data.get("rollegrupper", []) or []:
        if safe((gruppe.get("type") or {}).get("kode")) != "REGN":
            continue
        for rolle in gruppe.get("roller", []) or []:
            if rolle.get("fratraadt") is True or rolle.get("avregistrert") is True:
                continue
            enhet = rolle.get("enhet")
            if enhet:
                navn_list = enhet.get("navn", [])
                if navn_list:
                    return " ".join(navn_list).strip()
            person = rolle.get("person")
            if person:
                navn = build_full_name(person)
                if navn:
                    return navn
    return ""


def pick_primary_contact(candidates):
    for wanted in ["KONT", "DAGL", "STYR"]:
        for role_code, name in candidates:
            if role_code == wanted:
                return name, role_code
    return "", ""


def _extract_phone_from_html(html):
    """Extract phone number from Proff HTML (search or profile page)."""
    text = normalize_space(re.sub(r"<[^>]+>", " ", html))
    match = re.search(r"Telefon\s*([\d\s]{8,15})", text)
    if match:
        return normalize_space(match.group(1))
    return ""


def _extract_address_from_brreg(orgnr, timeout):
    """Fetch address from Brreg API (structured JSON)."""
    url = f"https://data.brreg.no/enhetsregisteret/api/enheter/{orgnr}"
    data = http_get_json(url, timeout=timeout)
    adr = data.get("forretningsadresse") or data.get("postadresse") or {}
    adresse_lines = adr.get("adresse", [])
    adresse = ", ".join(a for a in adresse_lines if a) if adresse_lines else ""
    postnr = safe(adr.get("postnummer"))
    poststed = safe(adr.get("poststed"))
    return adresse, postnr, poststed


def fetch_proff_phone(orgnr, timeout):
    """Fetch phone from Proff search page."""
    search_url = f"https://www.proff.no/bransjes%C3%B8k?q={orgnr}"
    html = http_get_text(search_url, timeout=timeout)

    phone = _extract_phone_from_html(html)
    if phone:
        return phone

    # Fallback: follow link to company profile page
    profile_match = re.search(r'href="(/selskap/[^"]+)"', html)
    if profile_match:
        profile_url = "https://www.proff.no" + profile_match.group(1)
        try:
            profile_html = http_get_text(profile_url, timeout=timeout)
            phone = _extract_phone_from_html(profile_html)
            if phone:
                return phone
        except Exception:
            pass

    return ""


def fetch_from_brreg(orgnr, timeout):
    url = f"https://data.brreg.no/enhetsregisteret/api/enheter/{orgnr}/roller"
    data = http_get_json(url, timeout=timeout)
    candidates = extract_brreg_candidates(data)
    regnskapsforer = extract_regnskapsforer(data)
    return pick_primary_contact(candidates), regnskapsforer


def normalize_space(text):
    return re.sub(r"\s+", " ", text).strip()


def _extract_styreleder_from_html(html):
    """Extract Styrets leder name from raw HTML using <a> link near the label."""
    link_match = re.search(
        r"styrets\s+leder\s*(?:<[^>]*>\s*)*<a[^>]*>\s*([^<]+?)\s*</a>",
        html,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if link_match:
        name = normalize_space(link_match.group(1))
        if name and len(name) > 2:
            return name

    # Fallback: strip tags and search plain text
    text = normalize_space(re.sub(r"<[^>]+>", " ", html))
    patterns = [
        r"Styrets leder\s+([A-ZÆØÅ][A-Za-zÆØÅæøåÉéÜüÖöÄä.\-\' ]+?)\s+(?:Adresse|\(f[\s\d])",
        r"Ledelse.administrasjon\s+Styrets leder\s+([A-ZÆØÅ][A-Za-zÆØÅæøåÉéÜüÖöÄä.\-\' ]+?)\s+\(",
        r"Styrets leder\s+([A-ZÆØÅ][A-Za-zÆØÅæøåÉéÜüÖöÄä.\-\' ]+?)\s+Kilde:\s*Brønnøysundregistrene",
        r"Styreleder\s+([A-ZÆØÅ][A-Za-zÆØÅæøåÉéÜüÖöÄä.\-\' ]+?)\s+Adresse",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            name = normalize_space(match.group(1))
            if name:
                return name

    return ""


def fetch_styreleder_from_proff(orgnr, timeout):
    search_url = f"https://www.proff.no/bransjes%C3%B8k?q={orgnr}"
    html = http_get_text(search_url, timeout=timeout)

    # Try extracting from search results page
    name = _extract_styreleder_from_html(html)
    if name:
        return name, "STYR", html

    # Fallback: follow link to company profile page
    profile_match = re.search(r'href="(/selskap/[^"]+)"', html)
    if profile_match:
        profile_url = "https://www.proff.no" + profile_match.group(1)
        try:
            profile_html = http_get_text(profile_url, timeout=timeout)
            name = _extract_styreleder_from_html(profile_html)
            if name:
                return name, "STYR", profile_html
        except Exception:
            pass

    return "", "", html


def fetch_primary_contact(orgnr, timeout):
    name, role, regnskapsforer = "", "", ""
    proff_html = None
    try:
        (name, role), regnskapsforer = fetch_from_brreg(orgnr, timeout)
    except Exception:
        pass

    if not name:
        try:
            name, role, proff_html = fetch_styreleder_from_proff(orgnr, timeout)
        except Exception:
            pass

    telefon, adresse, postnr, poststed = "", "", "", ""
    try:
        if proff_html:
            telefon = _extract_phone_from_html(proff_html)
        if not telefon:
            telefon = fetch_proff_phone(orgnr, timeout)
    except Exception:
        pass

    try:
        adresse, postnr, poststed = _extract_address_from_brreg(orgnr, timeout)
    except Exception:
        pass

    return name, role, telefon, adresse, postnr, poststed, regnskapsforer


def set_hyperlink(cell, text, url):
    cell.value = text
    cell.hyperlink = url
    cell.style = "Hyperlink"


def main():
    parser = argparse.ArgumentParser(
        description="Henter kontaktperson fra Brreg, med fallback til styreleder fra Proff."
    )
    parser.add_argument("--input", required=True, help="Input Excel-fil")
    parser.add_argument("--output", required=True, help="Output Excel-fil")
    parser.add_argument("--limit", type=int, default=None, help="Prosesser kun de første N radene")
    parser.add_argument("--timeout", type=int, default=DEFAULT_TIMEOUT, help="Timeout i sekunder")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"Fant ikke input-filen: {args.input}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(args.input)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if args.limit is not None:
        rows = rows[:args.limit]

    iterator = tqdm(rows, desc="Prosesserer") if tqdm is not None else rows

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Resultat"

    out_ws.append([
        "Selskap",
        "Organisasjonsnummer",
        "Kontaktperson navn",
        "Rolle",
        "Telefon",
        "Adresse",
        "Postnr",
        "Poststed",
        "Regnskapsfører",
        "Proff",
        "1881",
        "LinkedIn",
    ])
    out_ws.freeze_panes = "B2"

    total = len(rows)

    for i, row in enumerate(iterator, start=1):
        selskap = safe(row[0] if len(row) > 0 else "")
        orgnr = safe(row[1] if len(row) > 1 else "")

        navn, rolle, telefon, adresse, postnr, poststed, regnskapsforer = "", "", "", "", "", "", ""

        if orgnr:
            navn, rolle, telefon, adresse, postnr, poststed, regnskapsforer = fetch_primary_contact(orgnr, args.timeout)

        out_ws.append([selskap, orgnr, navn, rolle, telefon, adresse, postnr, poststed, regnskapsforer, "Proff", "1881", "LinkedIn"])
        r = out_ws.max_row

        if orgnr:
            set_hyperlink(
                out_ws.cell(r, 10),
                "Proff",
                f"https://www.proff.no/bransjes%C3%B8k?q={orgnr}",
            )

        if navn:
            q = quote_plus(navn)
            set_hyperlink(out_ws.cell(r, 11), "1881", f"https://www.1881.no/?query={q}")
            set_hyperlink(
                out_ws.cell(r, 12),
                "LinkedIn",
                f"https://www.linkedin.com/search/results/all/?keywords={q}",
            )

        if tqdm is None:
            print(f"{i}/{total} ferdig")

    out_wb.save(args.output)
    print("Ferdig:", args.output)
    os.startfile(args.output)


if __name__ == "__main__":
    main()
