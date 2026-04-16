import argparse
import os
import requests, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

API_BASE = "https://data.brreg.no/enhetsregisteret/api"

parser = argparse.ArgumentParser(description="Hent kundeliste fra Brønnøysundregistrene")
parser.add_argument("--input", required=True, help="Organisasjonsnummer å slå opp")
parser.add_argument("--output", default=None, help="Filnavn for XLSX-fil (standard: kunder_<orgnr>.xlsx)")
args = parser.parse_args()

ORG_NR = args.input
OUTPUT_FILE = args.output or f"kunder_{ORG_NR}.xlsx"

# Hent firmanavn
try:
    enhet_resp = requests.get(f"{API_BASE}/enheter/{ORG_NR}", headers={"Accept": "application/json"}, timeout=10)
    enhet_resp.raise_for_status()
    FIRMA_NAVN = enhet_resp.json().get("navn", ORG_NR)
except Exception:
    FIRMA_NAVN = ORG_NR

print(f"Firma: {FIRMA_NAVN} (org.nr {ORG_NR})")

# Hent alle sider
url = f"{API_BASE}/roller/enheter/{ORG_NR}/juridiskeroller"
alle_enheter = []
page = 1
while url:
    print(f"Henter side {page}...")
    resp = requests.get(url, headers={"Accept": "application/json"}, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    enheter = data.get("enheter", [])
    alle_enheter.extend(enheter)
    print(f"  {len(enheter)} enheter (totalt: {len(alle_enheter)})")
    url = data.get("_links", {}).get("next", {}).get("href")
    page += 1

print(f"\nTotalt hentet: {len(alle_enheter)} enheter")

kunder = []
for e in alle_enheter:
    navn = e.get("navn", "Ukjent")
    orgnr = e.get("organisasjonsnummer", "")
    roller = e.get("roller", [])
    aktive_roller = [r for r in roller if not r.get("fratraadt", False)]
    avregistrerte = [r for r in roller if r.get("avregistrert", False)]
    rolle_typer = ", ".join(sorted(set(r.get("type", {}).get("beskrivelse", "") for r in aktive_roller)))
    alle_rolle_typer = ", ".join(sorted(set(r.get("type", {}).get("beskrivelse", "") for r in roller)))
    fratraadt = len(aktive_roller) == 0
    avregistrert = len(avregistrerte) > 0 and len(avregistrerte) == len(roller)
    if avregistrert:
        status = "Avregistrert"
    elif fratraadt:
        status = "Fratrådt"
    else:
        status = "Aktiv"
    kunder.append({"navn": navn, "orgnr": orgnr, "rolle": rolle_typer if rolle_typer else alle_rolle_typer, "status": status})

kunder.sort(key=lambda x: (x["status"] != "Aktiv", x["navn"]))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Kunder"
hf = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
ha = Alignment(horizontal="center", vertical="center")
tb = Border(bottom=Side(style="thin", color="DDDDDD"))
af = Font(color="27AE60", bold=True)
rf = Font(color="E74C3C", bold=True)

for col, h in enumerate(["#", "Selskap", "Org.nr", "Rolle", "Status"], 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hf; c.fill = hfill; c.alignment = ha

for i, k in enumerate(kunder, 1):
    row = i + 1
    ws.cell(row=row, column=1, value=i).border = tb
    ws.cell(row=row, column=2, value=k["navn"]).border = tb
    ws.cell(row=row, column=3, value=k["orgnr"]).border = tb
    ws.cell(row=row, column=4, value=k["rolle"]).border = tb
    sc = ws.cell(row=row, column=5, value=k["status"])
    sc.border = tb
    sc.font = af if k["status"] == "Aktiv" else rf

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 45
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 16
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:E{len(kunder)+1}"

# Oppsummering
ws2 = wb.create_sheet("Oppsummering")
ws2.cell(row=1, column=1, value=f"{FIRMA_NAVN} - Kundeliste").font = Font(bold=True, size=14)
ws2.cell(row=2, column=1, value=f"Org.nr: {ORG_NR}")
ws2.cell(row=3, column=1, value="Generert: " + datetime.now().strftime("%Y-%m-%d %H:%M"))
ws2.cell(row=5, column=1, value="Statistikk").font = Font(bold=True, size=12)

totalt = len(kunder)
aktive = sum(1 for k in kunder if k["status"] == "Aktiv")
fratradte = sum(1 for k in kunder if k["status"] == "Fratrådt")
avreg = sum(1 for k in kunder if k["status"] == "Avregistrert")

for i, (lbl, val) in enumerate([("Totalt:", totalt), ("Aktive:", aktive), ("Fratrådte:", fratradte), ("Avregistrerte:", avreg)]):
    ws2.cell(row=6+i, column=1, value=lbl)
    ws2.cell(row=6+i, column=2, value=val).font = Font(bold=True)
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 10

out = OUTPUT_FILE
wb.save(out)
print(f"\nLagret: {out}")
print(f"  Totalt: {totalt}")
print(f"  Aktive: {aktive}")
print(f"  Fratrådte: {fratradte}")
print(f"  Avregistrerte: {avreg}")

print(f"\nÅpner {out}...")
os.startfile(out)
